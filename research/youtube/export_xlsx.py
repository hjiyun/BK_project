#!/usr/bin/env python3
"""채널 전체 영상 목록을 엑셀(xlsx)로 내보낸다.

열: 선택 | 채널명 | 업로드날짜 | 영상제목 | video_id | URL
'선택' 열에 O(또는 아무 값)를 표시한 행만 나중에 자막 수집 대상이 된다.
채널별 시트 + '전체' 시트 생성.
"""
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

BASE = Path(__file__).parent
MAN_DIR = BASE / "_channel_manifests"
OUT = BASE / "수집후보_영상목록.xlsx"

# 핸들 -> 표시용 채널명
NAMES = {
    "YeonmiParkOfficial": "박연미",
    "jooeuju-0815": "아오지누나",
    "yoonseolmi": "윤설미TV",
}

HEADERS = ["선택", "채널명", "업로드날짜", "영상제목", "video_id", "URL"]


def fmt_date(d):
    if not d or len(d) != 8:
        return "(날짜미상)"
    return f"{d[:4]}-{d[4:6]}-{d[6:]}"


def style_sheet(ws, nrows):
    hdr_fill = PatternFill("solid", fgColor="DDEBF7")
    for c, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True)
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center")
    widths = [6, 12, 13, 80, 14, 34]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{max(nrows, 1) + 1}"


def main():
    wb = Workbook()
    wb.remove(wb.active)

    all_rows = []
    for handle, name in NAMES.items():
        man = MAN_DIR / f"{handle}.jsonl"
        if not man.exists():
            print(f"[!] {man} 없음 — 건너뜀")
            continue
        recs = [json.loads(l) for l in man.open(encoding="utf-8") if l.strip()]
        # 최신순 유지(채널 /videos 순서), 날짜 있는 것 우선 정렬은 하지 않음
        rows = [["", name, fmt_date(r.get("upload_date")), r.get("title") or "(제목없음)",
                 r["id"], f"https://youtu.be/{r['id']}"] for r in recs]
        all_rows.extend(rows)

        ws = wb.create_sheet(title=name[:31])
        for r in rows:
            ws.append(r)
        style_sheet(ws, len(rows))
        print(f"[{name}] {len(rows)}행")

    # 전체 시트 (날짜 내림차순)
    ws = wb.create_sheet(title="전체", index=0)
    for r in sorted(all_rows, key=lambda x: x[2], reverse=True):
        ws.append(r)
    style_sheet(ws, len(all_rows))

    wb.save(OUT)
    print(f"\n저장: {OUT}  (총 {len(all_rows)}행)")


if __name__ == "__main__":
    main()
