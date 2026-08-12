#!/usr/bin/env python3
"""지금까지 수집한 자막 TXT 들의 메타를 모아 엑셀로 정리.

열: 채널 | 재생목록 | 업로드날짜 | 연도 | 영상제목 | video_id | 글자수 | URL
채널별 시트 + '전체' 시트.
"""
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

BASE = Path(__file__).parent
OUT = BASE / "수집완료_영상목록_2024.xlsx"

# (표시 채널명, 폴더). 실제 수집된 자막이 있는 곳만.
SOURCES = [
    ("유미카", "yumi_talkbox_2024"),
    ("주성하TV", "주성하TV_joosungha"),
    ("한송이", "한송이_songyi"),
    ("이철은", "이철은_Lee-CherUn"),
    ("강명도TV", "강명도TV_tv-6012"),
    ("강철환tv", "강철환tv"),
    ("배나TV", "배나TV_bnatv1004"),
    ("이소연TV", "이소연TV"),
]
HEADERS = ["채널", "재생목록", "업로드날짜", "연도", "영상제목", "video_id", "글자수", "URL"]


def parse(fp):
    """TXT 머리말 파싱 → dict. 머리말 형식이 소스마다 조금 달라 유연 처리."""
    txt = fp.read_text(encoding="utf-8", errors="replace")
    lines = txt.splitlines()
    title, meta = "", {}
    for ln in lines:
        if not ln.startswith("#"):
            break
        body = ln[1:].strip()
        m = re.match(r"([a-zA-Z_]+):\s*(.*)", body)
        if m and m.group(1).lower() in (
                "video_id", "upload_date", "url", "channel", "group",
                "playlist", "sub", "sub_lang"):
            meta[m.group(1).lower()] = m.group(2).strip()
        elif not title:
            title = body
    ud = meta.get("upload_date", "")
    return {
        "title": title,
        "vid": meta.get("video_id", ""),
        "ud": ud,
        "year": ud[:4] if ud else "",
        "url": meta.get("url", ""),
    }


def fmt_date(d):
    return f"{d[:4]}-{d[4:6]}-{d[6:]}" if len(d) == 8 else (d or "(미상)")


def style(ws, nrows):
    # 헤더는 반드시 데이터 append 전에 써야 함(안 그러면 첫 데이터행을 덮어씀)
    fill = PatternFill("solid", fgColor="DDEBF7")
    for c, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")
    for c, w in enumerate([12, 22, 13, 6, 78, 14, 8, 32], 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A2"


def finalize(ws, nrows):
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{max(nrows,1)+1}"


def main():
    wb = Workbook()
    wb.remove(wb.active)
    all_rows = []

    for chname, folder in SOURCES:
        root = BASE / folder
        if not root.exists():
            continue
        files = sorted(f for f in root.rglob("*.txt") if f.name != "_failed.txt")
        rows = []
        for f in files:
            m = parse(f)
            if m["year"] != "2024":          # 2024년만
                continue
            # 재생목록 = 폴더 구조상 채널 폴더 바로 아래 하위폴더명 (없으면 채널영상)
            rel = f.relative_to(root)
            playlist = rel.parts[0] if len(rel.parts) > 1 else "유미의 토크박스" if folder.startswith("yumi") else "(채널영상)"
            rows.append([chname, playlist, fmt_date(m["ud"]), m["year"],
                         m["title"], m["vid"], len(f.read_text(encoding="utf-8")), m["url"]])
        if not rows:
            continue
        all_rows.extend(rows)
        ws = wb.create_sheet(title=chname[:31])
        style(ws, len(rows))              # 헤더 먼저
        for r in rows:
            ws.append(list(r))
        finalize(ws, len(rows))
        print(f"[{chname}] {len(rows)}행")

    ws = wb.create_sheet(title="전체", index=0)
    style(ws, len(all_rows))              # 헤더 먼저
    for r in sorted(all_rows, key=lambda x: (x[0], x[2])):
        ws.append(list(r))
    finalize(ws, len(all_rows))

    wb.save(OUT)
    print(f"\n저장: {OUT}  (총 {len(all_rows)}행)")


if __name__ == "__main__":
    main()
